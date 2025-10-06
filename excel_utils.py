"""
Excel 파일 생성 유틸리티 모듈
openpyxl을 사용하여 Excel 파일을 생성하고 데이터를 입력합니다.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict, Any
import os

from invoice.models import InvoiceData
from packing_list.models import PackingListItem


def create_invoice_sheet(ws, invoices: List[InvoiceData]):
    """
    Invoice 시트를 생성합니다.
    
    Args:
        ws: 워크시트 객체
        invoices: InvoiceData 객체 리스트
    """
    # 헤더 정의 (수정된 필드명과 순서) - 파스칼 케이스 통일, ean → ref → ref00 순서
    headers = [
        'EDI', 'DeliveryNo', 'InvoiceNo', 'InvoiceDate', 
        'ShipmentNo', 'TotalQuantity',
        'EAN', 'Ref', 'Ref00', 'Description', 'Quantity', 'UnitPrice', 
        'TotalPriceUsd', 'Country'
    ]
    
    # 헤더 작성
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 데이터 작성
    row = 2
    for invoice in invoices:
        for item in invoice.items.values():
            # 날짜 포맷 변경 (dd.mm.yyyy -> yyyy-mm-dd)
            invoice_date = invoice.invoice_date
            if invoice_date and '.' in invoice_date:
                try:
                    day, month, year = invoice_date.split('.')
                    invoice_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                except:
                    pass  # 변환 실패시 원본 유지
            
            # REF_00 생성 (REF 끝 두자리를 00으로 변경)
            ref_00 = item.product_code
            if ref_00 and len(ref_00) >= 2:
                ref_00 = ref_00[:-2] + "00"
            
            # ShipmentNo에서 앞의 0000 제거
            clean_shipment_no = invoice.shipment_number
            if clean_shipment_no:
                clean_shipment_no = clean_shipment_no.lstrip('0')
                if not clean_shipment_no:  # 모든 문자가 0인 경우 '0'으로 설정
                    clean_shipment_no = '0'
            
            ws.cell(row=row, column=1, value=invoice.edi_number)
            ws.cell(row=row, column=2, value=invoice.delivery_number)
            ws.cell(row=row, column=3, value=invoice.invoice_number)
            ws.cell(row=row, column=4, value=invoice_date)
            ws.cell(row=row, column=5, value=clean_shipment_no)
            
            # 각 행의 TotalQuantity는 메타데이터에서 직접 사용
            try:
                ws.cell(row=row, column=6, value=int(invoice.total_quantity) if invoice.total_quantity else None)
            except:
                ws.cell(row=row, column=6, value=invoice.total_quantity)
            
            try:
                ws.cell(row=row, column=7, value=int(item.ean_number) if item.ean_number else None)
            except:
                ws.cell(row=row, column=7, value=item.ean_number)
            
            ws.cell(row=row, column=8, value=item.product_code)  # REF
            ws.cell(row=row, column=9, value=ref_00)  # REF_00
            ws.cell(row=row, column=10, value=item.description)
            
            try:
                ws.cell(row=row, column=11, value=int(item.quantity) if item.quantity else None)
            except:
                ws.cell(row=row, column=11, value=item.quantity)
            
            try:
                ws.cell(row=row, column=12, value=float(item.unit_price.replace(',', '')) if item.unit_price else None)
            except:
                ws.cell(row=row, column=12, value=item.unit_price)
            
            try:
                ws.cell(row=row, column=13, value=float(item.total_price_usd.replace(',', '')) if item.total_price_usd else None)
            except:
                ws.cell(row=row, column=13, value=item.total_price_usd)
            
            ws.cell(row=row, column=14, value=item.country)
            row += 1
    
    # 집계 테이블 추가 (Q열, 17번째 컬럼부터)
    _create_invoice_summary_table(ws, invoices, start_col=17)
    
    # 열 너비 자동 조정
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15


def _create_invoice_summary_table(ws, invoices: List[InvoiceData], start_col: int):
    """
    인보이스 집계 테이블을 생성합니다.
    
    Args:
        ws: 워크시트 객체
        invoices: InvoiceData 객체 리스트
        start_col: 시작 컬럼 번호 (Q열 = 17)
    """
    # 집계 테이블 헤더 (새로운 순서: ShipmentNo, InvoiceNo, InvoiceDate, TotalQuantity, TotalPriceUsd)
    summary_headers = ['ShipmentNo', 'InvoiceNo', 'InvoiceDate', 'TotalQuantity', 'TotalPriceUsd']
    
    # 헤더 작성
    for col, header in enumerate(summary_headers):
        cell = ws.cell(row=1, column=start_col + col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # 집계 데이터 작성
    current_row = 2
    total_quantity_sum = 0
    total_price_sum = 0
    
    for invoice in invoices:
        # 총 가격 계산
        total_price = 0
        for item in invoice.items.values():
            try:
                price = float(item.total_price_usd.replace(',', '')) if item.total_price_usd else 0
                total_price += price
            except:
                pass
        
        # ShipmentNo에서 앞의 0000 제거
        clean_shipment_no = invoice.shipment_number
        if clean_shipment_no:
            clean_shipment_no = clean_shipment_no.lstrip('0')
            if not clean_shipment_no:
                clean_shipment_no = '0'
        
        # 날짜 포맷 변경 (dd.mm.yyyy -> yyyy-mm-dd)
        invoice_date = invoice.invoice_date
        if invoice_date and '.' in invoice_date:
            try:
                day, month, year = invoice_date.split('.')
                invoice_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            except:
                pass
        
        ws.cell(row=current_row, column=start_col, value=clean_shipment_no)
        ws.cell(row=current_row, column=start_col + 1, value=invoice.invoice_number)
        ws.cell(row=current_row, column=start_col + 2, value=invoice_date)
        
        # 아이템들의 수량 합 계산 (메타데이터 대신)
        calculated_total_qty = 0
        for inv_item in invoice.items.values():
            try:
                qty = int(inv_item.quantity) if inv_item.quantity else 0
                calculated_total_qty += qty
            except:
                pass
        
        ws.cell(row=current_row, column=start_col + 3, value=calculated_total_qty)
        total_quantity_sum += calculated_total_qty
        
        ws.cell(row=current_row, column=start_col + 4, value=total_price)
        total_price_sum += total_price
        current_row += 1
    
    # 총합 행 추가
    total_cell = ws.cell(row=current_row, column=start_col, value="Total")
    total_cell.font = Font(bold=True)
    total_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 빈 셀들
    ws.cell(row=current_row, column=start_col + 1, value="")
    ws.cell(row=current_row, column=start_col + 2, value="")
    
    # 총합 값들
    qty_sum_cell = ws.cell(row=current_row, column=start_col + 3, value=total_quantity_sum)
    qty_sum_cell.font = Font(bold=True)
    qty_sum_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    price_sum_cell = ws.cell(row=current_row, column=start_col + 4, value=total_price_sum)
    price_sum_cell.font = Font(bold=True)
    price_sum_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 집계 테이블 열 너비 조정
    for col in range(5):
        ws.column_dimensions[chr(64 + start_col + col)].width = 15


def create_packing_list_sheet(ws, items: List[PackingListItem]):
    """
    Packing List 시트를 생성합니다.
    
    Args:
        ws: 워크시트 객체
        items: PackingListItem 객체 리스트
    """
    # 헤더 정의 (수정된 필드명과 순서) - EAN을 REF 앞으로 이동
    headers = [
        'EDI', 'DeliveryNo', 'ShipmentNo', 'Brand', 'EAN', 'REF', 'REF_00', 
        'Description', 'Qty', 'Batch', 'MfgDate', 'ExpDate', 'Dg'
    ]
    
    # 헤더 작성
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 데이터 작성
    for row, item in enumerate(items, 2):
        # REF_00 생성 (REF 끝 두자리를 00으로 변경)
        ref_00 = item.sku
        if ref_00 and len(ref_00) >= 2:
            ref_00 = ref_00[:-2] + "00"
        
        # 날짜 포맷 변경 (dd-mm-yyyy -> yyyy-mm-dd)
        mfg_date = item.mfg_date
        if mfg_date and '-' in mfg_date:
            try:
                day, month, year = mfg_date.split('-')
                mfg_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            except:
                pass  # 변환 실패시 원본 유지
        
        exp_date = item.exp_date
        if exp_date and '-' in exp_date:
            try:
                day, month, year = exp_date.split('-')
                exp_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            except:
                pass  # 변환 실패시 원본 유지
        
        # DeliveryNo에서 앞의 00 제거
        clean_delivery_no = item.order_number
        if clean_delivery_no and clean_delivery_no.startswith('00'):
            clean_delivery_no = clean_delivery_no[2:]  # 앞 2자리 제거
            if not clean_delivery_no:  # 모든 문자가 0인 경우 '0'으로 설정
                clean_delivery_no = '0'
        
        # ShipmentNo에서 앞의 4자리만 제거 (0000)
        clean_shipment_no = item.shipment_number
        if clean_shipment_no and len(clean_shipment_no) >= 4 and clean_shipment_no.startswith('0000'):
            clean_shipment_no = clean_shipment_no[4:]  # 앞 4자리 제거
            if not clean_shipment_no:  # 모든 문자가 0인 경우 '0'으로 설정
                clean_shipment_no = '0'
        
        ws.cell(row=row, column=1, value=item.edi_number)
        ws.cell(row=row, column=2, value=clean_delivery_no)
        ws.cell(row=row, column=3, value=clean_shipment_no)
        ws.cell(row=row, column=4, value=item.brand)
        
        # EAN을 5번째 컬럼으로 이동 (숫자 타입으로 변환)
        try:
            ws.cell(row=row, column=5, value=int(item.ean) if item.ean else None)
        except:
            ws.cell(row=row, column=5, value=item.ean)
        
        ws.cell(row=row, column=6, value=item.sku)  # REF
        ws.cell(row=row, column=7, value=ref_00)  # REF_00
        ws.cell(row=row, column=8, value=item.description)
        
        try:
            ws.cell(row=row, column=9, value=int(item.items_qty) if item.items_qty else None)
        except:
            ws.cell(row=row, column=9, value=item.items_qty)
        
        ws.cell(row=row, column=10, value=item.batch)
        ws.cell(row=row, column=11, value=mfg_date)
        ws.cell(row=row, column=12, value=exp_date)
        ws.cell(row=row, column=13, value=item.dg)
    
    # 집계 테이블 추가 (O열, 15번째 컬럼부터)
    _create_packing_list_summary_table(ws, items, start_col=15)
    
    # 열 너비 자동 조정
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15


def _create_packing_list_summary_table(ws, items: List[PackingListItem], start_col: int):
    """
    패킹리스트 집계 테이블을 생성합니다.
    
    Args:
        ws: 워크시트 객체
        items: PackingListItem 객체 리스트
        start_col: 시작 컬럼 번호 (O열 = 15)
    """
    # shipment_number별 items_qty 집계
    shipment_summary = {}
    total_sum = 0
    
    for item in items:
        shipment_no = item.shipment_number
        if shipment_no:
            # ShipmentNo에서 앞의 4자리만 제거 (0000) - 집계 테이블용
            clean_shipment_no = shipment_no
            if shipment_no and len(shipment_no) >= 4 and shipment_no.startswith('0000'):
                clean_shipment_no = shipment_no[4:]  # 앞 4자리 제거
                if not clean_shipment_no:  # 모든 문자가 0인 경우 '0'으로 설정
                    clean_shipment_no = '0'
            
            if clean_shipment_no not in shipment_summary:
                shipment_summary[clean_shipment_no] = 0
            try:
                qty = int(item.items_qty) if item.items_qty else 0
                shipment_summary[clean_shipment_no] += qty
                total_sum += qty
            except:
                pass
    
    # 집계 테이블 헤더
    summary_headers = ['ShipmentNo', 'TotalQty']
    
    # 헤더 작성
    for col, header in enumerate(summary_headers):
        cell = ws.cell(row=1, column=start_col + col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # 집계 데이터 작성
    current_row = 2
    for shipment_no, total_qty in shipment_summary.items():
        ws.cell(row=current_row, column=start_col, value=shipment_no)
        ws.cell(row=current_row, column=start_col + 1, value=total_qty)
        current_row += 1
    
    # 총합 행 추가
    total_cell = ws.cell(row=current_row, column=start_col, value="Total")
    total_cell.font = Font(bold=True)
    total_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    sum_cell = ws.cell(row=current_row, column=start_col + 1, value=total_sum)
    sum_cell.font = Font(bold=True)
    sum_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 집계 테이블 열 너비 조정
    for col in range(2):
        ws.column_dimensions[chr(64 + start_col + col)].width = 15


def create_structured_excel(
    output_path: str, 
    invoices: List[InvoiceData] = None, 
    packing_items: List[PackingListItem] = None
):
    """
    구조화된 데이터로 Excel 파일을 생성합니다.
    
    Args:
        output_path: 출력할 Excel 파일 경로
        invoices: Invoice 데이터 리스트
        packing_items: Packing List 아이템 리스트
    
    Raises:
        ValueError: 데이터가 모두 None일 때
        Exception: Excel 파일 생성 오류 시
    """
    if not invoices and not packing_items:
        raise ValueError("Invoice 또는 Packing List 데이터 중 최소 하나는 제공되어야 합니다.")
    
    try:
        # 새 워크북 생성
        wb = Workbook()
        
        # 기본 시트 제거
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Invoice 시트 생성
        if invoices:
            ws_invoice = wb.create_sheet(title="Invoice")
            create_invoice_sheet(ws_invoice, invoices)
        
        # Packing List 시트 생성
        if packing_items:
            ws_packing = wb.create_sheet(title="Packing_List")
            create_packing_list_sheet(ws_packing, packing_items)
        
        # 파일 저장
        wb.save(output_path)
        print(f"Excel 파일이 생성되었습니다: {output_path}")
        
    except Exception as e:
        raise Exception(f"Excel 파일 생성 오류: {str(e)}")


# Legacy functions for backward compatibility
def create_excel_with_sheets(output_path, sheets_data):
    """레거시 함수 - 하위 호환성을 위해 유지"""
    try:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        for sheet_name, text_data in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            ws['A1'] = text_data
            ws['A1'].font = Font(name='맑은 고딕', size=10)
            ws['A1'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.column_dimensions['A'].width = min(100, max(20, len(text_data) // 50))
        
        wb.save(output_path)
    except Exception as e:
        raise Exception(f"Excel 파일 생성 오류: {str(e)}")


def write_to_excel(output_path, pl_text=None, ci_text=None):
    """레거시 함수 - 하위 호환성을 위해 유지"""
    if not pl_text and not ci_text:
        raise ValueError("PL 또는 CI 텍스트 중 최소 하나는 제공되어야 합니다.")
    
    sheets_data = {}
    if pl_text:
        sheets_data["PL"] = pl_text
    if ci_text:
        sheets_data["CI"] = ci_text
    
    create_excel_with_sheets(output_path, sheets_data)


def get_output_directory(pdf_path):
    """PDF 파일이 위치한 디렉토리 경로를 반환합니다."""
    return os.path.dirname(os.path.abspath(pdf_path))
